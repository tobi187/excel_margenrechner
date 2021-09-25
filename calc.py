from openpyxl import load_workbook


def calc_marge(file_name, product_name, amount):

    wb = load_workbook(file_name, data_only=True)
    row_nr = -1
    marge = 0
    for sheet in wb.worksheets:
        for row in range(1, 100000):
            if sheet.cell(row=row, column=1).value == product_name:
                row_nr = row
                break

        marge = sheet.cell(row=row_nr, column=3) * amount - sheet.cell(row=row_nr, column=2) * amount - sheet.cell(
            row=row_nr, column=4) * amount - sheet.cell(row=row_nr, column=5) * amount - sheet.cell(
            row=row_nr, column=6) * amount

    return marge


